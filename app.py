"""
퍼블리싱 코드 자동 검수 도구 - Flask 웹 서버
"""
import io
import os
import re
import json
import shutil
import subprocess
import traceback
import zipfile
import tempfile
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from src.reviewer import Reviewer

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 200 * 1024 * 1024  # 최대 200MB

# 가장 최근 검수 파일 내용 보관 (AI 수정 제안 시 컨텍스트 제공용)
# {filepath: content} — 단일 세션만 유지 (메모리 절약)
_last_review_files: dict = {}


@app.route('/')
def index():
    """메인 페이지"""
    return render_template('index.html')


@app.route('/review', methods=['POST'])
def review():
    """검수 실행 API"""
    if 'file' not in request.files:
        return jsonify({'error': '파일이 없습니다'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '파일을 선택해주세요'}), 400

    if not file.filename.endswith('.zip'):
        return jsonify({'error': 'ZIP 파일만 업로드 가능합니다'}), 400

    # 선택된 규칙
    selected_rules = request.form.getlist('rules')
    if not selected_rules:
        selected_rules = [
            'typography', 'color', 'important', 'scrollbar',
            'gradient', 'style_mixing', 'console_log', 'zindex', 'accessibility',
        ]

    # 임시 폴더에 저장 후 검수
    with tempfile.TemporaryDirectory() as tmp_dir:
        zip_path = os.path.join(tmp_dir, 'upload.zip')
        file.save(zip_path)

        try:
            reviewer = Reviewer(zip_path, selected_rules)
            report = reviewer.run()

            # 파일 내용을 세션에 보관 (AI 수정 제안 컨텍스트용)
            global _last_review_files
            _last_review_files = reviewer.get_files()

            return jsonify(report)
        except zipfile.BadZipFile:
            return jsonify({'error': '올바른 ZIP 파일이 아닙니다'}), 400
        except MemoryError:
            return jsonify({'error': 'ZIP 파일이 너무 큽니다. node_modules·dist·build 폴더를 제외하고 소스 코드만 압축해주세요.'}), 400
        except Exception as e:
            traceback.print_exc()
            return jsonify({'error': f'검수 중 오류 발생: {str(e)}'}), 500


def _get_file_context(filepath: str, target_line: int, radius: int = 10) -> str:
    """저장된 파일에서 target_line 전후 radius줄 컨텍스트를 추출합니다."""
    content = _last_review_files.get(filepath, '')
    if not content or target_line <= 0:
        return ''
    lines = content.splitlines()
    start = max(0, target_line - 1 - radius)
    end = min(len(lines), target_line + radius)
    result = []
    for i, line in enumerate(lines[start:end], start=start + 1):
        marker = '▶' if i == target_line else ' '
        result.append(f"L.{i:4d} {marker} {line}")
    return '\n'.join(result)


def _find_claude_cli():
    """Claude Code CLI 실행 경로 찾기"""
    for name in ('claude', 'claude.cmd'):
        found = shutil.which(name)
        if found:
            return found
    for env_var in ('APPDATA', 'LOCALAPPDATA'):
        base = os.environ.get(env_var, '')
        for name in ('claude.cmd', 'claude'):
            c = os.path.join(base, 'npm', name)
            if os.path.isfile(c):
                return c
    return None


def _call_claude(prompt: str) -> str | None:
    """Claude Code CLI를 호출해서 분석 결과를 받아옴"""
    found = _find_claude_cli()
    if not found:
        return None
    try:
        r = subprocess.run(
            [found, '-p', '-', '--dangerously-skip-permissions'],
            input=prompt.encode('utf-8'),
            capture_output=True,
            timeout=150,
        )
        if r.returncode == 0:
            return r.stdout.decode('utf-8', errors='replace').strip()
    except Exception:
        pass
    return None


@app.route('/api/ai-summary', methods=['POST'])
def ai_summary():
    """검수 결과를 AI가 분석해서 코멘트 생성"""
    data = request.get_json()
    summary = data.get('summary', {})
    issues = data.get('issues', [])

    critical_issues = [i for i in issues if i.get('severity') == 'critical']
    warning_issues = [i for i in issues if i.get('severity') == 'warning']

    # 카테고리별 대표 이슈 정리 (너무 길면 Claude 응답 느려짐)
    issue_lines = []
    for issue in issues[:20]:
        line = f"- [{issue.get('severity','?')}] {issue.get('category','')} | {issue.get('file','')} | {issue.get('description','')}"
        issue_lines.append(line)

    prompt = f"""당신은 VCP(Vehicle Content Platform) 퍼블리싱 코드 검수 전문가입니다.
현대오토에버의 퍼블리싱 코드 가이드라인 기준으로 아래 검수 결과를 분석해주세요.

【검수 결과 요약】
- 검사 파일 수: {summary.get('total_files', 0)}개
- 심각한 문제: {summary.get('critical', 0)}건
- 주의 사항: {summary.get('warnings', 0)}건
- 카테고리별: {json.dumps(summary.get('by_category', {}), ensure_ascii=False)}

【주요 이슈 목록】
{chr(10).join(issue_lines) if issue_lines else '이슈 없음'}

비개발자(기획자)도 이해할 수 있게 쉬운 말로 설명해주세요.
아래 JSON 형식으로만 응답하세요:
{{
  "overall": "전체 코드 품질에 대한 2-3줄 요약. 심각도와 개선 필요성을 포함.",
  "bullets": ["핵심 이슈 1", "핵심 이슈 2", "핵심 이슈 3"],
  "action": "개발팀에 전달할 조치 요청 사항. 1-2줄."
}}"""

    text = _call_claude(prompt)
    if not text:
        return jsonify({'error': 'Claude CLI를 찾을 수 없습니다'}), 500

    match = re.search(r'\{[\s\S]*\}', text)
    if match:
        try:
            result = json.loads(match.group())
            return jsonify({'summary': result})
        except json.JSONDecodeError:
            pass

    return jsonify({'summary': {'overall': text, 'bullets': [], 'action': ''}})


@app.route('/api/suggest-fix', methods=['POST'])
def suggest_fix():
    """이슈의 실제 코드를 기반으로 Claude가 구체적인 수정 방법 제안"""
    data = request.get_json()
    category  = data.get('category', '')
    file_path = data.get('file', '')
    description = data.get('description', '')
    occurrences = data.get('occurrences', [])   # [{line, code}, ...]
    reason    = data.get('reason', '')

    # 발견된 위반 코드 목록 (최대 15줄)
    code_lines = []
    for o in occurrences[:15]:
        file_label = f"[{o['file']}] " if o.get('file') else ''
        line_label = f"L.{o['line']}  " if o.get('line') else ''
        code_lines.append(f"{file_label}{line_label}{o.get('code', '')}")
    code_block = '\n'.join(code_lines) if code_lines else '(코드 정보 없음)'

    # 파일 원문 ±10줄 컨텍스트 (첫 번째 occurrence 기준)
    first_line = occurrences[0].get('line', 0) if occurrences else 0
    file_context = _get_file_context(file_path, first_line, radius=10)
    context_section = (
        f'\n【파일 원문 컨텍스트 (L.{first_line} 전후 10줄, ▶ 표시가 위반 라인)】\n{file_context}'
        if file_context else ''
    )

    # 카테고리별 가이드라인 컨텍스트 (HMG 매핑 포함, 단호한 어조)
    guidelines = {
        '타이포그래피': (
            '【위반 기준】 fontSize / fontWeight / lineHeight 등의 값을 숫자로 직접 입력하는 것은\n'
            '현대오토에버 VCP 코드 가이드라인 위반입니다. 반드시 HMG 디자인 시스템 토큰을 사용해야 합니다.\n\n'
            '【올바른 방식】\n'
            '  - MUI sx prop: sx={{ ...theme.typography.body2 }}\n'
            '  - 또는: import { typography } from \'@/shared/theme\';\n'
            '    sx={{ ...typography.body2 }}\n\n'
            '【주요 타이포그래피 토큰】\n'
            '  typography.h1 / h2 / h3 / h4 / h5 / h6\n'
            '  typography.subtitle1 / subtitle2\n'
            '  typography.body1 / body2\n'
            '  typography.caption / overline / button'
        ),
        '색상': (
            '【위반 기준】 색상 코드(#hex, rgb(), rgba())를 코드에 직접 입력하는 것은\n'
            '현대오토에버 VCP 코드 가이드라인 위반입니다. 반드시 HMG Design System 색상 변수를 사용해야 합니다.\n\n'
            '【올바른 방식】\n'
            '  import { colors } from \'@/shared/theme\';\n'
            '  sx={{ color: colors.gray900, backgroundColor: colors.white }}\n\n'
            '【HMG 색상 변수 매핑표 - 반드시 이 변수명을 사용할 것】\n'
            '  #000000          → colors.black\n'
            '  #111111          → colors.gray900\n'
            '  #333333          → colors.gray800\n'
            '  #666666          → colors.gray600\n'
            '  #999999          → colors.gray400\n'
            '  #cccccc, #CCCCCC → colors.gray200\n'
            '  #e9eaec, #E9EAEC → colors.gray100\n'
            '  #ffffff, #FFFFFF → colors.white\n'
            '  #8333e6, #8333E6 → colors.purple500 (또는 colors.primary)\n'
            '  위 목록에 없는 색상 → 프로젝트 내 색상 상수 파일에 정의 후 참조'
        ),
        '!important': (
            '【위반 기준】 CSS !important 남용은 현대오토에버 VCP 코드 가이드라인 위반입니다.\n'
            '한 줄에 !important를 2개 이상 쓰는 것은 즉시 수정이 필요한 심각한 문제입니다.\n\n'
            '【올바른 방식】\n'
            '  방법1. CSS 선택자 명시도(specificity) 높이기:\n'
            '    .parent .child .target { color: red; }  /* !important 없이 */\n'
            '  방법2. MUI sx prop 우선순위 활용:\n'
            '    sx={{ \'& .MuiButton-root\': { color: colors.primary } }}\n'
            '  방법3. 꼭 필요한 경우 1개로 제한, 한 줄에 여러 !important 절대 금지'
        ),
        '스크롤바': (
            '【위반 기준】 ::-webkit-scrollbar 만 단독 사용 시 Firefox, 일부 Edge에서 동작하지 않습니다.\n'
            '현대오토에버 VCP 코드 가이드라인: 반드시 브라우저 호환 스타일을 함께 작성해야 합니다.\n\n'
            '【올바른 방식】\n'
            '  방법1. 표준 속성 병행 작성 (두 방식 모두 작성):\n'
            '    /* 표준 (Firefox, 최신 브라우저) */\n'
            '    scrollbar-width: thin;\n'
            '    scrollbar-color: #888888 transparent;\n'
            '    /* Webkit (Chrome, Safari) */\n'
            '    &::-webkit-scrollbar { width: 6px; }\n'
            '    &::-webkit-scrollbar-thumb { background: #888888; }\n'
            '  방법2. @/shared/theme의 scrollbarSx 유틸리티 재사용:\n'
            '    import { scrollbarSx } from \'@/shared/theme\';\n'
            '    sx={{ ...scrollbarSx }}'
        ),
        '그라데이션': (
            '【위반 기준】 linear-gradient, radial-gradient 값을 직접 하드코딩하는 것은\n'
            '현대오토에버 VCP 코드 가이드라인 위반입니다. 동일한 그라데이션이 여러 곳에 반복되면 안 됩니다.\n\n'
            '【올바른 방식】\n'
            '  방법1. 프로젝트 공통 constants/theme에 정의 후 참조:\n'
            '    // constants/gradients.ts\n'
            '    export const GRADIENT_MAIN = \'linear-gradient(135deg, #8333e6 0%, #111111 100%)\';\n'
            '    // 사용처\n'
            '    import { GRADIENT_MAIN } from \'@/shared/constants/gradients\';\n'
            '    sx={{ background: GRADIENT_MAIN }}\n'
            '  방법2. CSS 변수로 정의:\n'
            '    --gradient-primary: linear-gradient(...);\n'
            '    background: var(--gradient-primary);'
        ),
        '스타일링 통일': (
            '【위반 기준】 MUI sx prop과 인라인 style={{}}, makeStyles/withStyles가 혼용되면\n'
            '현대오토에버 VCP 코드 가이드라인 위반입니다.\n\n'
            '【올바른 방식 - MUI v5 표준】\n'
            '  방법1. sx prop:\n'
            '    <Box sx={{ color: colors.gray900, fontSize: 14 }} />\n'
            '  방법2. styled() 컴포넌트:\n'
            '    const StyledBox = styled(Box)(({ theme }) => ({\n'
            '      color: theme.palette.primary.main,\n'
            '    }));\n\n'
            '  makeStyles/withStyles는 MUI v4 방식으로 v5에서 deprecated.\n'
            '  인라인 style={{}} 은 동적 값이 꼭 필요한 경우에만 최소한으로 사용.'
        ),
        'console.log': (
            '【위반 기준】 console.log, console.error, console.warn, debugger 등이\n'
            '배포 코드에 잔류하는 것은 현대오토에버 VCP 코드 기준 위반입니다.\n\n'
            '【조치 방법】\n'
            '  - console.log 줄 전체를 삭제하세요.\n'
            '  - 운영 환경 로그가 필요한 경우: 프로젝트 공통 logger 유틸리티 사용\n'
            '  - debugger는 반드시 삭제. 배포 환경에서 브라우저를 멈추게 합니다.'
        ),
        'z-index': (
            '【위반 기준】 z-index에 999, 9999 같은 임의의 큰 숫자를 직접 입력하면\n'
            '현대오토에버 VCP 코드 기준 위반입니다.\n\n'
            '【MUI zIndex 토큰 매핑표 - 반드시 이 토큰을 사용할 것】\n'
            '  theme.zIndex.mobileStepper = 1000\n'
            '  theme.zIndex.fab           = 1050\n'
            '  theme.zIndex.appBar        = 1200\n'
            '  theme.zIndex.drawer        = 1300\n'
            '  theme.zIndex.modal         = 1400\n'
            '  theme.zIndex.snackbar      = 1500\n'
            '  theme.zIndex.tooltip       = 1600\n\n'
            '【올바른 방식】\n'
            '  sx={{ zIndex: theme.zIndex.modal }}\n'
            '  또는 styled(Box)(({ theme }) => ({ zIndex: theme.zIndex.appBar }))'
        ),
        '접근성': (
            '【위반 기준】 outline: none, img alt 누락, button type 누락은\n'
            '현대오토에버 VCP 코드 기준 및 WCAG 2.1 접근성 가이드라인 위반입니다.\n\n'
            '【outline: none 대체 방법】\n'
            '  /* 키보드 포커스일 때만 보이게 */\n'
            '  &:focus-visible {\n'
            '    outline: 2px solid colors.primary;\n'
            '    outline-offset: 2px;\n'
            '  }\n\n'
            '【img alt 필수 작성】\n'
            '  <img src={src} alt="차량 사이드 뷰 이미지" />  /* 의미있는 이미지 */\n'
            '  <img src={src} alt="" role="presentation" />   /* 장식용 이미지 */\n\n'
            '【button type 명시】\n'
            '  <button type="button">클릭</button>\n'
            '  <button type="submit">제출</button>'
        ),
    }
    guideline = guidelines.get(category, '현대오토에버 VCP 퍼블리싱 코드 가이드라인을 준수하세요.')

    prompt = f"""당신은 현대오토에버 VCP 퍼블리싱 코드 검수 전문가입니다.
아래 코드에서 기준 위반 사항을 발견했습니다. 명확하고 단호하게 수정 방법을 안내해주세요.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{guideline}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

【문제 분류】 {category}
【파일】 {file_path}
【문제 설명】 {description}
【문제 이유】 {reason}

【발견된 위반 코드 목록】
{code_block}
{context_section}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
다음 형식으로 답변해주세요 (마크다운 코드블록 없이 plain text):

■ 무엇이 문제인가
위 실제 코드에서 발견된 구체적인 위반 내용. "이 코드의 OOO은 위반입니다"처럼 직접적으로. (2-3줄)

■ 수정 전
위의 발견된 실제 코드에서 대표적인 1-3줄을 그대로 인용. (새로 만들지 말 것)

■ 수정 후
위 코드를 가이드라인의 올바른 방식으로 직접 수정한 실제 코드.
색상의 경우 반드시 위 매핑표의 변수명 사용. 임의로 변수명을 만들지 말 것.

■ 참고사항
추가 주의사항이 있으면 1-2줄. 없으면 이 항목 자체를 쓰지 말 것."""

    text = _call_claude(prompt)
    if not text:
        return jsonify({'error': 'Claude CLI를 찾을 수 없습니다'}), 500

    return jsonify({'suggestion': text})


CATEGORY_GUIDELINES = {
    '타이포그래피': (
        '【위반】 fontSize/fontWeight/lineHeight 값을 숫자로 직접 입력. '
        '【수정】 HMG 디자인 시스템 타이포그래피 토큰 사용: '
        'import { typography } from \'@/shared/theme\'; → sx={{ ...typography.body2 }}'
    ),
    '색상': (
        '【위반】 #hex 색상 코드를 직접 입력. '
        '【수정】 HMG Design System 색상 변수 사용: '
        'import { colors } from \'@/shared/theme\'; → sx={{ color: colors.gray900 }}. '
        '매핑: #111111→gray900, #333333→gray800, #666666→gray600, '
        '#999999→gray400, #cccccc→gray200, #ffffff→white, #8333e6→purple500(primary)'
    ),
    '!important': (
        '【위반】 CSS !important 과다 사용. 한 줄에 2개 이상은 즉시 수정 필요. '
        '【수정】 CSS 선택자 명시도(specificity) 높이기 또는 MUI sx prop 우선순위 조정으로 제거.'
    ),
    '스크롤바': (
        '【위반】 ::-webkit-scrollbar 단독 사용 (Chrome/Safari 전용). '
        '【수정】 scrollbar-width: thin; scrollbar-color: ... 표준 속성 병행 작성 '
        '또는 @/shared/theme의 scrollbarSx 유틸리티 재사용.'
    ),
    '그라데이션': (
        '【위반】 linear-gradient/radial-gradient 값을 직접 하드코딩. '
        '【수정】 HMG 디자인 시스템 그라데이션 변수 또는 프로젝트 공통 constants에 정의 후 참조.'
    ),
    '스타일링 통일': (
        '【위반】 MUI sx prop과 인라인 style={{}}, makeStyles/withStyles 혼용. '
        '【수정】 MUI v5 표준인 sx prop 또는 styled() 컴포넌트로 통일.'
    ),
    'console.log': (
        '【위반】 console.log / debugger 배포 코드 잔류. '
        '【수정】 해당 줄 완전 삭제. 운영 로그 필요 시 프로젝트 공통 logger 유틸리티 사용.'
    ),
    'z-index': (
        '【위반】 z-index에 임의의 큰 숫자(100 이상) 직접 사용. '
        '【수정】 MUI theme.zIndex 토큰 사용: '
        'appBar(1200), drawer(1300), modal(1400), snackbar(1500), tooltip(1600).'
    ),
    '접근성': (
        '【위반】 outline:none으로 키보드 포커스 제거, img alt 누락, button type 누락. '
        '【수정】 outline: none → :focus-visible 대체; '
        'img에 alt 속성 필수; button에 type="button"|"submit" 명시.'
    ),
}


def _batch_fix_suggestions(issues: list) -> list[str]:
    """모든 이슈에 대해 Claude에게 일괄 수정 제안 요청 → 순서대로 문자열 리스트 반환"""
    if not issues:
        return []

    issue_blocks = []
    for i, iss in enumerate(issues, start=1):
        occs = iss.get('occurrences', [])[:5]
        code_sample = '\n'.join(
            f"  L.{o.get('line','')}  {o.get('code','')}"
            for o in occs
        ) or '  (코드 샘플 없음)'
        guide = CATEGORY_GUIDELINES.get(iss.get('category', ''), '')
        issue_blocks.append(
            f"[{i}] 카테고리: {iss.get('category','')} | 심각도: {iss.get('severity','')}\n"
            f"    파일: {iss.get('file','')}\n"
            f"    설명: {iss.get('description','')}\n"
            f"    가이드라인: {guide}\n"
            f"    실제 코드:\n{code_sample}"
        )

    prompt = (
        "당신은 VCP 퍼블리싱 코드 전문가입니다.\n"
        "아래 이슈 목록 각각에 대해 구체적인 수정 방법을 작성해주세요.\n\n"
        "반드시 아래 JSON 배열 형식으로만 응답하세요 (다른 텍스트 없이):\n"
        '[\n  {"fix": "이슈1 수정방법"},\n  {"fix": "이슈2 수정방법"},\n  ...\n]\n\n'
        "각 fix는 다음을 포함해야 합니다:\n"
        "- 문제가 되는 코드 패턴 (수정 전)\n"
        "- 올바른 코드 예시 (수정 후)\n"
        "- 간단한 설명 (1-2줄)\n\n"
        "이슈 목록:\n\n"
        + '\n\n'.join(issue_blocks)
    )

    text = _call_claude(prompt)
    if not text:
        return ['(AI 제안 생성 실패)'] * len(issues)

    match = re.search(r'\[[\s\S]*\]', text)
    if not match:
        return ['(AI 응답 파싱 실패)'] * len(issues)

    try:
        parsed = json.loads(match.group())
        result = []
        for i, item in enumerate(parsed):
            if i >= len(issues):
                break
            result.append(item.get('fix', ''))
        # 부족하면 빈 값으로 채움
        while len(result) < len(issues):
            result.append('')
        return result
    except json.JSONDecodeError:
        return ['(JSON 파싱 오류)'] * len(issues)


@app.route('/api/download-excel', methods=['POST'])
def download_excel():
    """검수 결과를 AI 수정 제안 포함 상세 엑셀로 다운로드"""
    data = request.get_json()
    summary  = data.get('summary', {})
    issues   = data.get('issues', [])
    filename = data.get('filename', 'VCP_퍼블리싱')
    today    = datetime.now().strftime('%Y-%m-%d')
    today_kr = datetime.now().strftime('%Y년 %m월 %d일')

    # ── 1. AI 수정 제안 일괄 생성 ─────────────────────────────────
    sorted_issues = sorted(
        issues,
        key=lambda x: (0 if x.get('severity') == 'critical' else 1, x.get('file', ''))
    )
    fix_suggestions = _batch_fix_suggestions(sorted_issues)

    # ── 2. 공통 스타일 헬퍼 ──────────────────────────────────────
    def bd(color='BBBBBB', style='thin'):
        s = Side(border_style=style, color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    def bd_outer(color='888888'):
        s = Side(border_style='medium', color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    def fill(hex_color):
        return PatternFill('solid', fgColor=hex_color)

    C = {
        'navy':    '1A2744',
        'navy2':   '2D3F6B',
        'gray1':   'F5F5F5',
        'gray2':   'E8E8E8',
        'gray3':   '9E9E9E',
        'red_bg':  'FFF0F0',
        'red_txt': 'C0392B',
        'org_bg':  'FFF8EC',
        'org_txt': 'B7560A',
        'grn_bg':  'F0FFF4',
        'grn_txt': '1A6636',
        'white':   'FFFFFF',
        'blk':     '1A1A1A',
    }

    def font(size=9, bold=False, color='1A1A1A', name='맑은 고딕'):
        return Font(name=name, size=size, bold=bold, color=color)

    def align(h='center', v='center', wrap=True):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    wb = Workbook()

    # ════════════════════════════════════════════════════════════
    # Sheet 1: 검수 결과 요약
    # ════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = '검수 결과 요약'
    ws1.sheet_view.showGridLines = False

    for col, w in zip('ABCDEF', [28, 16, 16, 16, 16, 24]):
        ws1.column_dimensions[col].width = w

    # 타이틀 블록
    ws1.row_dimensions[1].height = 46
    ws1.merge_cells('A1:F1')
    c = ws1['A1']
    c.value = 'VCP 퍼블리싱 코드 검수 결과 보고서'
    c.font  = font(18, True, C['white'])
    c.fill  = fill(C['navy'])
    c.alignment = align()

    ws1.row_dimensions[2].height = 18
    ws1.merge_cells('A2:F2')
    c = ws1['A2']
    c.value = f'검수 파일: {filename}    검수일: {today_kr}    작성: INNOCEAN'
    c.font  = font(9, False, 'AAAAAA')
    c.fill  = fill(C['navy2'])
    c.alignment = align()

    ws1.row_dimensions[3].height = 10  # 여백

    # 핵심 지표 4개 (가로 배치)
    ws1.row_dimensions[4].height = 20
    ws1.row_dimensions[5].height = 38
    ws1.row_dimensions[6].height = 20

    metric_configs = [
        ('검사 파일 수',     summary.get('total_files', 0),    '개', C['navy'],   C['white']),
        ('수정 필수',        summary.get('critical', 0),       '건', C['red_txt'], C['red_bg']),
        ('개선 권고',        summary.get('warnings', 0),       '건', C['org_txt'], C['org_bg']),
        ('총 이슈',          summary.get('total_issues', 0),   '건', C['navy2'],  'EEF2FF'),
    ]
    metric_cols = ['A', 'B', 'C', 'D']
    for col_l, (label, val, unit, txt_c, bg_c) in zip(metric_cols, metric_configs):
        ws1[f'{col_l}4'].value = label
        ws1[f'{col_l}4'].font  = font(9, True, C['gray3'])
        ws1[f'{col_l}4'].fill  = fill(bg_c)
        ws1[f'{col_l}4'].alignment = align()
        ws1[f'{col_l}4'].border = bd()

        ws1[f'{col_l}5'].value = f'{val} {unit}'
        ws1[f'{col_l}5'].font  = font(20, True, txt_c)
        ws1[f'{col_l}5'].fill  = fill(bg_c)
        ws1[f'{col_l}5'].alignment = align()
        ws1[f'{col_l}5'].border = bd()

        ws1[f'{col_l}6'].value = ''
        ws1[f'{col_l}6'].fill  = fill(bg_c)
        ws1[f'{col_l}6'].border = bd()

    ws1.row_dimensions[7].height = 14  # 여백

    # 카테고리별 현황 표
    ws1.row_dimensions[8].height = 22
    for col_l, label in zip(['A', 'B', 'C'], ['카테고리', '건수', '비율']):
        c = ws1[f'{col_l}8']
        c.value = label
        c.font  = font(9, True, C['white'])
        c.fill  = fill(C['navy'])
        c.alignment = align()
        c.border = bd(C['navy'])

    total_issues = summary.get('total_issues', 1) or 1
    for r_off, (cat, cnt) in enumerate(summary.get('by_category', {}).items(), start=9):
        ws1.row_dimensions[r_off].height = 20
        pct = f'{cnt/total_issues*100:.1f}%'
        for col_l, val in zip(['A', 'B', 'C'], [cat, cnt, pct]):
            c = ws1[f'{col_l}{r_off}']
            c.value = val
            c.font  = font(9)
            c.fill  = fill(C['gray1'])
            c.alignment = align()
            c.border = bd()

    # 파일별 현황 표
    cat_end_row = 8 + len(summary.get('by_category', {}))
    sep_row = cat_end_row + 2

    ws1.row_dimensions[sep_row].height = 22
    for col_l, label in zip(['A', 'B', 'C', 'D', 'E'],
                             ['파일명', '수정 필수', '개선 권고', '합계', '주요 카테고리']):
        c = ws1[f'{col_l}{sep_row}']
        c.value = label
        c.font  = font(9, True, C['white'])
        c.fill  = fill(C['navy2'])
        c.alignment = align()
        c.border = bd(C['navy2'])

    file_stats: dict = {}
    for iss in issues:
        f = iss.get('file', '(미상)')
        if f not in file_stats:
            file_stats[f] = {'critical': 0, 'warning': 0, 'cats': set()}
        file_stats[f][iss.get('severity', 'warning')] += 1
        file_stats[f]['cats'].add(iss.get('category', ''))

    file_rows_sorted = sorted(
        file_stats.items(),
        key=lambda x: -(x[1]['critical'] * 10 + x[1]['warning'])
    )
    for r_off, (fname, stat) in enumerate(file_rows_sorted, start=sep_row + 1):
        ws1.row_dimensions[r_off].height = 18
        total = stat['critical'] + stat['warning']
        row_bg = C['red_bg'] if stat['critical'] > 0 else C['org_bg']
        for col_l, val in zip(
            ['A', 'B', 'C', 'D', 'E'],
            [fname, stat['critical'] or '', stat['warning'] or '',
             total, ', '.join(stat['cats'])]
        ):
            c = ws1[f'{col_l}{r_off}']
            c.value = val
            c.font  = font(9, col_l == 'B' and stat['critical'] > 0, C['blk'])
            if col_l == 'B' and stat['critical'] > 0:
                c.font = font(9, True, C['red_txt'])
            c.fill  = fill(row_bg)
            c.alignment = align('left' if col_l in ('A', 'E') else 'center')
            c.border = bd()

    # ════════════════════════════════════════════════════════════
    # Sheet 2: 이슈 상세 목록 (AI 수정 제안 포함)
    # ════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet('이슈 상세 목록')
    ws2.sheet_view.showGridLines = False

    col_specs = [
        ('No.',         5),
        ('심각도',       10),
        ('카테고리',      12),
        ('파일 경로',     46),
        ('줄번호',        7),
        ('이슈 내용',     40),
        ('위반 코드 (상위 5건)', 52),
        ('위반 건수',     8),
        ('AI 수정 제안',  55),
    ]
    for i, (name, w) in enumerate(col_specs, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
        c = ws2.cell(row=1, column=i, value=name)
        c.font      = font(9, True, C['white'])
        c.fill      = fill(C['navy'])
        c.alignment = align()
        c.border    = bd(C['navy'])

    ws2.row_dimensions[1].height = 24
    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = f'A1:{get_column_letter(len(col_specs))}1'

    for idx, (issue, fix) in enumerate(zip(sorted_issues, fix_suggestions), start=1):
        sev     = issue.get('severity', '')
        row_bg  = C['red_bg'] if sev == 'critical' else C['org_bg']
        sev_lbl = '수정 필수' if sev == 'critical' else '개선 권고'
        sev_clr = C['red_txt'] if sev == 'critical' else C['org_txt']

        # 위반 코드 샘플 (상위 5줄)
        occs = issue.get('occurrences', [])[:5]
        code_sample = '\n'.join(
            (f"[{o['file']}] " if o.get('file') else '') +
            (f"L.{o['line']}  " if o.get('line') else '') +
            o.get('code', '')
            for o in occs
        )
        total_occs = len(issue.get('occurrences', []))

        row_num = idx + 1
        # 행 높이: fix 텍스트 길이에 비례
        line_count = max(
            len(str(fix).split('\n')),
            len(code_sample.split('\n')),
            3
        )
        ws2.row_dimensions[row_num].height = max(60, line_count * 14)

        values_aligns = [
            (idx,                        align()),
            (sev_lbl,                    align()),
            (issue.get('category', ''),  align()),
            (issue.get('file', ''),      align('left')),
            (issue.get('line', '') or '', align()),
            (issue.get('description',''), align('left')),
            (code_sample,                align('left')),
            (total_occs if total_occs else '', align()),
            (fix,                        align('left')),
        ]

        for col_i, (val, aln) in enumerate(values_aligns, start=1):
            c = ws2.cell(row=row_num, column=col_i, value=val)
            if col_i == 2:  # 심각도 컬럼 색상
                c.font = font(9, True, sev_clr)
            elif col_i == 7:  # 위반 코드 - monospace
                c.font = Font(name='Courier New', size=8, color=C['blk'])
            elif col_i == 9:  # AI 수정 제안
                c.font = font(9, False, '2D3F6B')
            else:
                c.font = font(9)
            c.fill      = fill(row_bg)
            c.alignment = aln
            c.border    = bd()

    # ════════════════════════════════════════════════════════════
    # Sheet 3: 파일별 이슈 분포
    # ════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet('파일별 이슈 분포')
    ws3.sheet_view.showGridLines = False

    ws3.row_dimensions[1].height = 32
    ws3.merge_cells('A1:G1')
    c = ws3['A1']
    c.value = '파일별 이슈 분포'
    c.font  = font(13, True, C['white'])
    c.fill  = fill(C['navy'])
    c.alignment = align()

    all_cats = sorted(summary.get('by_category', {}).keys())
    hdr3 = ['파일명', '수정 필수', '개선 권고', '합계'] + all_cats
    ws3.column_dimensions['A'].width = 52
    for i in range(1, len(hdr3)):
        ws3.column_dimensions[get_column_letter(i + 1)].width = 13

    ws3.row_dimensions[2].height = 22
    for i, h in enumerate(hdr3, start=1):
        c = ws3.cell(row=2, column=i, value=h)
        c.font      = font(9, True, C['white'])
        c.fill      = fill(C['navy2'])
        c.alignment = align()
        c.border    = bd(C['navy2'])

    ws3.freeze_panes = 'A3'
    ws3.auto_filter.ref = f'A2:{get_column_letter(len(hdr3))}2'

    # 파일별 카테고리 교차 집계
    file_cat: dict = {}
    for iss in issues:
        f   = iss.get('file', '(미상)')
        cat = iss.get('category', '')
        sev = iss.get('severity', 'warning')
        if f not in file_cat:
            file_cat[f] = {'critical': 0, 'warning': 0}
            for cat_ in all_cats:
                file_cat[f][cat_] = 0
        file_cat[f]['critical' if sev == 'critical' else 'warning'] += 1
        if cat in file_cat[f]:
            file_cat[f][cat] += 1

    for r_off, (fname, stat) in enumerate(
        sorted(file_cat.items(), key=lambda x: -(x[1]['critical'] * 100 + x[1]['warning'])),
        start=3
    ):
        ws3.row_dimensions[r_off].height = 18
        total = stat['critical'] + stat['warning']
        row_bg = C['red_bg'] if stat['critical'] > 0 else (C['org_bg'] if total > 0 else C['gray1'])

        row_vals = [fname, stat['critical'] or '', stat['warning'] or '', total] + \
                   [stat.get(cat, '') or '' for cat in all_cats]
        for col_i, val in enumerate(row_vals, start=1):
            c = ws3.cell(row=r_off, column=col_i, value=val)
            c.font = font(9, col_i == 2 and stat['critical'] > 0,
                          C['red_txt'] if col_i == 2 and stat['critical'] > 0 else C['blk'])
            c.fill      = fill(row_bg)
            c.alignment = align('left' if col_i == 1 else 'center')
            c.border    = bd()

    # ── 저장 및 반환 ─────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = re.sub(r'[^\w가-힣]', '_', filename)
    dl_name   = f'VCP검수결과_{safe_name}_{today}.xlsx'

    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=dl_name,
    )


@app.route('/api/draft-mail', methods=['POST'])
def draft_mail():
    """원더무브 발송용 메일 초안을 Claude가 작성"""
    data = request.get_json()
    summary = data.get('summary', {})
    issues = data.get('issues', [])
    filename = data.get('filename', '')
    today = datetime.now().strftime('%Y년 %m월 %d일')

    # 카테고리별 요약
    by_cat = summary.get('by_category', {})
    issue_lines = []
    for issue in issues[:30]:
        sev = '【수정 필수】' if issue.get('severity') == 'critical' else '【개선 권고】'
        issue_lines.append(
            f"{sev} [{issue.get('category','')}] {issue.get('file','')} "
            f"{'L.' + str(issue.get('line')) if issue.get('line') else ''} "
            f"- {issue.get('description','')}"
        )

    prompt = f"""당신은 INNOCEAN 소속 기획자입니다.
원더무브(퍼블리싱 외주사)에게 보내는 공식 업무 이메일을 작성해주세요.

【배경】
- INNOCEAN은 원더무브로부터 VCP 퍼블리싱 코드를 납품받아 현대오토에버에 전달합니다.
- 이번 검수에서 아래와 같은 코드 품질 문제가 발견되었습니다.
- 현대오토에버가 제시한 검수 기준({', '.join(by_cat.keys())})에 어긋나는 항목입니다.

【검수 결과】
- 대상 파일: {filename}
- 검수일: {today}
- 심각한 문제(수정 필수): {summary.get('critical', 0)}건
- 개선 권고: {summary.get('warnings', 0)}건
- 카테고리: {json.dumps(by_cat, ensure_ascii=False)}

【발견된 주요 이슈 (상위 {min(len(issue_lines), 30)}건)】
{chr(10).join(issue_lines)}

【이메일 작성 조건】
1. 한국 대기업 공문 스타일 (정중하고 격식 있는 표현)
2. 수신: 원더무브 담당자  발신: INNOCEAN 담당자 (이름은 [담당자명] 자리 표시)
3. 제목 포함
4. 발견된 이슈를 카테고리별로 표로 정리 (파일명, 줄번호, 내용)
5. 수정 요청 + 수정본 전달 일정 회신 요청 포함
6. Outlook에 바로 붙여넣기 가능한 텍스트 형식 (HTML 태그 없이)
7. 이메일 본문만 출력 (설명/마크다운 없이)"""

    text = _call_claude(prompt)
    if not text:
        return jsonify({'error': 'Claude CLI를 찾을 수 없습니다'}), 500

    return jsonify({'mail': text})


@app.route('/api/ai-deep-review', methods=['POST'])
def ai_deep_review():
    """이슈 상위 파일을 AI가 직접 읽고 규칙 엔진이 놓친 추가 이슈 발견"""
    data = request.get_json()
    issues = data.get('issues', [])

    # 이슈 많은 상위 3개 파일 선택
    file_issue_count: dict = {}
    for issue in issues:
        f = issue.get('file', '')
        if f:
            file_issue_count[f] = file_issue_count.get(f, 0) + 1

    top_files = sorted(file_issue_count, key=file_issue_count.get, reverse=True)[:3]  # type: ignore[arg-type]

    # 이슈가 없으면 보관된 파일 중 상위 3개
    if not top_files:
        top_files = list(_last_review_files.keys())[:3]

    if not top_files or not _last_review_files:
        return jsonify({'error': '분석할 파일이 없습니다. 먼저 검수를 실행해주세요.'}), 400

    results = []
    for filepath in top_files:
        content = _last_review_files.get(filepath, '')
        if not content:
            continue

        # 파일이 너무 길면 앞 150줄만 전달
        lines = content.splitlines()
        truncated = len(lines) > 150
        snippet = '\n'.join(lines[:150]) if truncated else content
        trunc_note = f'\n(이하 {len(lines) - 150}줄 생략)' if truncated else ''

        prompt = f"""당신은 현대자동차그룹 VCP(Vehicle Content Platform) 퍼블리싱 코드 품질 검수 전문가입니다.
아래 파일을 직접 읽고, 기존 자동 검수 도구가 놓쳤을 수 있는 추가적인 코드 품질 문제를 찾아주세요.

【파일명】 {filepath}
【소스코드】
{snippet}{trunc_note}

【검수 기준 (현대오토에버 VCP 기준)】
- HMG 디자인 시스템 color 토큰 미사용 (#hex 직접 사용 → colors.gray900 등으로 교체)
- MUI theme.typography 미사용 (font-size/font-weight/line-height 하드코딩)
- outline: none / outline: 0 (키보드 접근성 파괴)
- z-index 매직넘버 100 이상 직접 사용 (→ theme.zIndex 토큰)
- console.log / debugger 잔류 (배포 코드)
- !important 남용 (한 줄에 2개 이상 또는 반복적 사용)
- 인라인 style={{}} 과 sx prop 혼용
- gradient 하드코딩 (linear-gradient 등)
- <img> alt 속성 누락
- <button> type 속성 누락
- 그 외 VCP 기준 위반이나 명백한 코드 품질 문제

【출력 형식】
발견된 추가 이슈를 아래 JSON 배열로만 출력하세요 (설명 없이):
[
  {{
    "line": 줄번호(숫자),
    "severity": "critical" 또는 "warning",
    "category": "카테고리명",
    "description": "한 줄 설명",
    "code": "해당 코드 (한 줄)",
    "suggestion": "수정 방법 한 줄"
  }}
]

이슈가 없으면 빈 배열 []을 출력하세요. JSON 외 다른 텍스트는 절대 출력하지 마세요."""

        text = _call_claude(prompt)
        if not text:
            continue

        try:
            json_match = re.search(r'\[.*\]', text, re.DOTALL)
            if json_match:
                found = json.loads(json_match.group())
                if isinstance(found, list) and found:
                    results.append({
                        'file': filepath,
                        'issues': found[:10],  # 파일당 최대 10개
                    })
        except (json.JSONDecodeError, ValueError):
            continue

    return jsonify({'findings': results})


if __name__ == '__main__':
    print("퍼블리싱 코드 검수 도구를 시작합니다...")
    print("브라우저에서 http://localhost:5000 을 열어주세요")
    app.run(debug=True, port=5000)
